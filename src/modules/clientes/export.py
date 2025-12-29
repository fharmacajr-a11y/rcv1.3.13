"""Exportação de clientes para formatos CSV e Excel (opcional).

Este módulo fornece funções headless para exportar dados de clientes
em formatos CSV (padrão) e XLSX (se openpyxl estiver disponível).
"""

from __future__ import annotations

import csv
import logging
from pathlib import Path
from typing import TYPE_CHECKING, Any, Dict, List

if TYPE_CHECKING:
    from .viewmodel import ClienteRow

logger = logging.getLogger(__name__)


# Definir colunas de exportação
CSV_COLUMNS = [
    "id",
    "razao_social",
    "cnpj",
    "nome",
    "whatsapp",
    "observacoes",
    "status",
    "ultima_alteracao",
]

# Cabeçalhos amigáveis
CSV_HEADERS = {
    "id": "ID",
    "razao_social": "Razão Social",
    "cnpj": "CNPJ",
    "nome": "Nome Fantasia",
    "whatsapp": "WhatsApp",
    "observacoes": "Observações",
    "status": "Status",
    "ultima_alteracao": "Última Alteração",
}


def export_clients_to_csv(rows: List[ClienteRow], output_path: Path) -> None:
    """Exporta clientes para arquivo CSV.

    Args:
        rows: Lista de ClienteRow a serem exportados.
        output_path: Caminho completo do arquivo CSV de destino.

    Raises:
        IOError: Se houver erro ao escrever o arquivo.

    Note:
        Usa encoding 'utf-8-sig' para compatibilidade com Excel em PT-BR.
        O BOM (Byte Order Mark) garante que o Excel interprete corretamente
        caracteres acentuados.
    """
    logger.info("Exportando %d cliente(s) para CSV: %s", len(rows), output_path)

    try:
        with open(output_path, "w", encoding="utf-8-sig", newline="") as csvfile:
            # Usar DictWriter com cabeçalhos amigáveis
            writer = csv.DictWriter(
                csvfile,
                fieldnames=CSV_COLUMNS,
                extrasaction="ignore",  # Ignora campos extras
            )

            # Escrever cabeçalho
            writer.writerow(CSV_HEADERS)

            # Escrever dados
            for row in rows:
                writer.writerow(_row_to_dict(row))

        logger.info("Exportação CSV concluída: %s", output_path)

    except Exception as exc:
        logger.error("Erro ao exportar CSV: %s", exc)
        raise IOError(f"Falha ao exportar CSV: {exc}") from exc


def export_clients_to_xlsx(rows: List[ClienteRow], output_path: Path) -> None:
    """Exporta clientes para arquivo Excel (XLSX).

    Args:
        rows: Lista de ClienteRow a serem exportados.
        output_path: Caminho completo do arquivo XLSX de destino.

    Raises:
        ImportError: Se openpyxl não estiver instalado.
        IOError: Se houver erro ao escrever o arquivo.

    Note:
        Requer openpyxl instalado. Se não disponível, use export_clients_to_csv.
    """
    try:
        from openpyxl import Workbook  # pyright: ignore[reportMissingModuleSource]
        from openpyxl.styles import Font  # pyright: ignore[reportMissingModuleSource]
    except ImportError as exc:
        logger.error("openpyxl não está instalado")
        raise ImportError("Exportação XLSX requer openpyxl. Instale com: pip install openpyxl") from exc

    logger.info("Exportando %d cliente(s) para XLSX: %s", len(rows), output_path)

    try:
        # Criar workbook
        wb = Workbook()
        ws = wb.active
        if ws:
            ws.title = "Clientes"

            # Escrever cabeçalho com formatação
            header_font = Font(bold=True)
            for col_idx, col_name in enumerate(CSV_COLUMNS, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = CSV_HEADERS[col_name]
                cell.font = header_font

            # Escrever dados
            for row_idx, row in enumerate(rows, start=2):
                row_dict = _row_to_dict(row)
                for col_idx, col_name in enumerate(CSV_COLUMNS, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=row_dict.get(col_name, ""))

            # Ajustar largura das colunas (estimativa básica)
            for col_idx, col_name in enumerate(CSV_COLUMNS, start=1):
                # Largura baseada no tamanho do cabeçalho
                col_letter = ws.cell(row=1, column=col_idx).column_letter
                ws.column_dimensions[col_letter].width = max(len(CSV_HEADERS[col_name]) + 2, 12)

        # Salvar arquivo
        wb.save(output_path)
        logger.info("Exportação XLSX concluída: %s", output_path)

    except ImportError:
        raise  # Re-raise ImportError
    except Exception as exc:
        logger.error("Erro ao exportar XLSX: %s", exc)
        raise IOError(f"Falha ao exportar XLSX: {exc}") from exc


def _row_to_dict(row: ClienteRow) -> Dict[str, Any]:
    """Converte ClienteRow para dicionário com colunas de exportação.

    Args:
        row: Objeto ClienteRow.

    Returns:
        Dicionário com valores das colunas definidas em CSV_COLUMNS.
    """
    return {
        "id": row.id,
        "razao_social": row.razao_social,
        "cnpj": row.cnpj,
        "nome": row.nome,
        "whatsapp": row.whatsapp,
        "observacoes": row.observacoes,
        "status": row.status,
        "ultima_alteracao": row.ultima_alteracao,
    }


def is_xlsx_available() -> bool:
    """Verifica se openpyxl está disponível para exportação XLSX.

    Returns:
        True se openpyxl pode ser importado, False caso contrário.
    """
    try:
        import openpyxl  # noqa: F401  # pyright: ignore[reportMissingModuleSource]

        return True
    except ImportError:
        return False


__all__ = [
    "export_clients_to_csv",
    "export_clients_to_xlsx",
    "is_xlsx_available",
    "CSV_COLUMNS",
    "CSV_HEADERS",
]
