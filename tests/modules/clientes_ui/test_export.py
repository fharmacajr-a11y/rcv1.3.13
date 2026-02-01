# -*- coding: utf-8 -*-
"""Testes para exportação CSV/Excel do ClientesV2 (FASE 3.5).

Valida:
- Exportação CSV com cabeçalho e linhas corretas
- Exportação XLSX usando openpyxl
- Export respeita filtros/busca (dados visíveis)
- Cancelamento do dialog não exporta
"""

import csv
import pytest
from unittest.mock import MagicMock, patch
from src.modules.clientes.ui.view import ClientesV2Frame
from src.modules.clientes.core.viewmodel import ClienteRow


@pytest.fixture
def mock_vm():
    """Mock do ClientesViewModel."""
    with patch("src.modules.clientes.ui.view.ClientesViewModel") as mock:
        vm_instance = MagicMock()
        mock.return_value = vm_instance
        yield vm_instance


@pytest.fixture
def sample_clients():
    """Dados de exemplo para testes."""
    return [
        ClienteRow(
            id=1,
            razao_social="Empresa Alpha LTDA",
            cnpj="11.111.111/0001-11",
            nome="Alpha",
            whatsapp="11987654321",
            status="Ativo",
            observacoes="Cliente VIP",
            ultima_alteracao="2024-01-15",
        ),
        ClienteRow(
            id=2,
            razao_social="Beta Comércio SA",
            cnpj="22.222.222/0001-22",
            nome="Beta",
            whatsapp="21987654321",
            status="Inativo",
            observacoes="",
            ultima_alteracao="2024-02-10",
        ),
        ClienteRow(
            id=3,
            razao_social="Gamma Tech",
            cnpj="33.333.333/0001-33",
            nome="Gamma",
            whatsapp="31987654321",
            status="Ativo",
            observacoes="Novo cliente",
            ultima_alteracao="2024-03-20",
        ),
    ]


@patch("tkinter.filedialog.asksaveasfilename")
def test_export_csv_with_header_and_rows(mock_dialog, root, mock_vm, sample_clients, tmp_path):
    """FASE 3.5: Exporta CSV com cabeçalho e linhas corretas."""
    frame = ClientesV2Frame(root)

    # Simular dados na tree
    for client in sample_clients:
        iid = f"I{client.id:03d}"
        frame._row_data_map[iid] = client
        frame.tree.insert(
            "",
            "end",
            iid=iid,
            values=(
                client.id,
                client.razao_social,
                client.cnpj,
                client.nome,
                client.whatsapp,
                client.status,
                client.observacoes,
                client.ultima_alteracao,
            ),
        )

    # Configurar path de saída
    output_file = tmp_path / "test_export.csv"
    mock_dialog.return_value = str(output_file)

    # Chamar export
    frame._on_export()

    # Verificar que arquivo foi criado
    assert output_file.exists(), "Arquivo CSV deve ser criado"

    # Validar conteúdo do CSV
    with open(output_file, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    # Verificar número de linhas
    assert len(rows) == 3, "Deve exportar 3 clientes"

    # Verificar cabeçalhos
    assert rows[0].keys() == {
        "ID",
        "Razão Social",
        "CNPJ",
        "Nome Fantasia",
        "WhatsApp",
        "Observações",
        "Status",
        "Última Alteração",
    }

    # Verificar primeira linha
    assert rows[0]["ID"] == "1"
    assert rows[0]["Razão Social"] == "Empresa Alpha LTDA"
    assert rows[0]["CNPJ"] == "11.111.111/0001-11"
    assert rows[0]["Status"] == "Ativo"

    # Verificar segunda linha
    assert rows[1]["ID"] == "2"
    assert rows[1]["Razão Social"] == "Beta Comércio SA"


@patch("tkinter.filedialog.asksaveasfilename")
def test_export_xlsx_creates_valid_file(mock_dialog, root, mock_vm, sample_clients, tmp_path):
    """FASE 3.5: Exporta XLSX e valida com openpyxl."""
    pytest.importorskip("openpyxl", reason="openpyxl não disponível")

    from openpyxl import load_workbook

    frame = ClientesV2Frame(root)

    # Simular dados na tree
    for client in sample_clients:
        iid = f"I{client.id:03d}"
        frame._row_data_map[iid] = client
        frame.tree.insert(
            "",
            "end",
            iid=iid,
            values=(
                client.id,
                client.razao_social,
                client.cnpj,
                client.nome,
                client.whatsapp,
                client.status,
                client.observacoes,
                client.ultima_alteracao,
            ),
        )

    # Configurar path de saída XLSX
    output_file = tmp_path / "test_export.xlsx"
    mock_dialog.return_value = str(output_file)

    # Chamar export
    frame._on_export()

    # Verificar que arquivo foi criado
    assert output_file.exists(), "Arquivo XLSX deve ser criado"

    # Validar usando openpyxl
    wb = load_workbook(output_file)
    assert len(wb.sheetnames) >= 1, "Deve ter pelo menos 1 sheet"

    ws = wb.active
    assert ws is not None, "Sheet ativa deve existir"

    # Verificar cabeçalhos (linha 1)
    assert ws.cell(1, 1).value == "ID"
    assert ws.cell(1, 2).value == "Razão Social"
    assert ws.cell(1, 3).value == "CNPJ"

    # Verificar primeira linha de dados (linha 2)
    assert ws.cell(2, 1).value == 1
    assert ws.cell(2, 2).value == "Empresa Alpha LTDA"
    assert ws.cell(2, 3).value == "11.111.111/0001-11"

    # Verificar total de linhas (header + 3 clientes)
    assert ws.max_row == 4, "Deve ter 4 linhas (1 header + 3 dados)"


@patch("tkinter.filedialog.asksaveasfilename")
def test_export_respects_filters(mock_dialog, root, mock_vm, sample_clients, tmp_path):
    """FASE 3.5: Export usa apenas dados visíveis/filtrados."""
    frame = ClientesV2Frame(root)

    # Simular apenas 2 clientes filtrados (removendo o terceiro)
    filtered_clients = sample_clients[:2]  # Apenas Alpha e Beta

    for client in filtered_clients:
        iid = f"I{client.id:03d}"
        frame._row_data_map[iid] = client
        frame.tree.insert(
            "",
            "end",
            iid=iid,
            values=(
                client.id,
                client.razao_social,
                client.cnpj,
                client.nome,
                client.whatsapp,
                client.status,
                client.observacoes,
                client.ultima_alteracao,
            ),
        )

    # Configurar path de saída
    output_file = tmp_path / "test_filtered.csv"
    mock_dialog.return_value = str(output_file)

    # Chamar export
    frame._on_export()

    # Verificar que exportou apenas os 2 filtrados
    with open(output_file, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    assert len(rows) == 2, "Deve exportar apenas 2 clientes filtrados"
    assert rows[0]["Razão Social"] == "Empresa Alpha LTDA"
    assert rows[1]["Razão Social"] == "Beta Comércio SA"

    # Verificar que Gamma NÃO foi exportado
    assert not any(r["Razão Social"] == "Gamma Tech" for r in rows)


@patch("tkinter.filedialog.asksaveasfilename")
def test_export_cancel_does_not_create_file(mock_dialog, root, mock_vm, sample_clients, tmp_path):
    """FASE 3.5: Cancelar dialog não cria arquivo."""
    frame = ClientesV2Frame(root)

    # Simular dados
    for client in sample_clients:
        iid = f"I{client.id:03d}"
        frame._row_data_map[iid] = client
        frame.tree.insert(
            "",
            "end",
            iid=iid,
            values=(
                client.id,
                client.razao_social,
                client.cnpj,
                client.nome,
                client.whatsapp,
                client.status,
                client.observacoes,
                client.ultima_alteracao,
            ),
        )

    # Simular cancelamento (dialog retorna string vazia)
    mock_dialog.return_value = ""

    # Chamar export
    frame._on_export()

    # Verificar que nenhum arquivo CSV/XLSX foi criado em tmp_path
    csv_files = list(tmp_path.glob("*.csv"))
    xlsx_files = list(tmp_path.glob("*.xlsx"))
    assert len(csv_files) == 0, "Nenhum arquivo CSV deve ser criado ao cancelar"
    assert len(xlsx_files) == 0, "Nenhum arquivo XLSX deve ser criado ao cancelar"


@patch("tkinter.filedialog.asksaveasfilename")
@patch("tkinter.messagebox.showinfo")
def test_export_empty_data_shows_message(mock_msgbox, mock_dialog, root, mock_vm):
    """FASE 3.5: Export sem dados mostra mensagem."""
    frame = ClientesV2Frame(root)

    # Não adicionar dados (tree vazia)

    # Chamar export
    frame._on_export()

    # Verificar que messagebox foi chamado
    assert mock_msgbox.called, "Deve mostrar mensagem quando não há dados"

    # Verificar que dialog NÃO foi aberto
    assert not mock_dialog.called, "Dialog não deve abrir sem dados"


@patch("tkinter.filedialog.asksaveasfilename")
def test_export_csv_encoding_utf8_sig(mock_dialog, root, mock_vm, tmp_path):
    """FASE 3.5: CSV usa encoding utf-8-sig para compatibilidade Excel."""
    frame = ClientesV2Frame(root)

    # Cliente com caracteres acentuados
    client = ClienteRow(
        id=99,
        razao_social="Açúcar & Café LTDA",
        cnpj="99.999.999/0001-99",
        nome="Açúcar",
        whatsapp="11999999999",
        status="Ativo",
        observacoes="Órgão público",
        ultima_alteracao="2024-12-31",
    )

    iid = "I099"
    frame._row_data_map[iid] = client
    frame.tree.insert(
        "",
        "end",
        iid=iid,
        values=(
            client.id,
            client.razao_social,
            client.cnpj,
            client.nome,
            client.whatsapp,
            client.status,
            client.observacoes,
            client.ultima_alteracao,
        ),
    )

    # Configurar path de saída
    output_file = tmp_path / "test_encoding.csv"
    mock_dialog.return_value = str(output_file)

    # Chamar export
    frame._on_export()

    # Verificar BOM (Byte Order Mark) no início do arquivo
    with open(output_file, "rb") as f:
        first_bytes = f.read(3)

    # UTF-8-SIG começa com BOM: EF BB BF
    assert first_bytes == b"\xef\xbb\xbf", "Arquivo deve ter BOM UTF-8"

    # Verificar que caracteres acentuados foram preservados
    with open(output_file, "r", encoding="utf-8-sig") as f:
        content = f.read()

    assert "Açúcar & Café LTDA" in content
    assert "Órgão público" in content


@patch("tkinter.filedialog.asksaveasfilename")
def test_export_handles_special_characters(mock_dialog, root, mock_vm, tmp_path):
    """FASE 3.5: Export lida corretamente com caracteres especiais."""
    frame = ClientesV2Frame(root)

    # Cliente com vírgulas, aspas e quebras de linha
    client = ClienteRow(
        id=1,
        razao_social='Empresa "Teste", LTDA',
        cnpj="11.111.111/0001-11",
        nome="Teste",
        whatsapp="11987654321",
        status="Ativo",
        observacoes="Linha 1\nLinha 2",
        ultima_alteracao="2024-01-01",
    )

    iid = "I001"
    frame._row_data_map[iid] = client
    frame.tree.insert(
        "",
        "end",
        iid=iid,
        values=(
            client.id,
            client.razao_social,
            client.cnpj,
            client.nome,
            client.whatsapp,
            client.status,
            client.observacoes,
            client.ultima_alteracao,
        ),
    )

    # Configurar path de saída
    output_file = tmp_path / "test_special.csv"
    mock_dialog.return_value = str(output_file)

    # Chamar export
    frame._on_export()

    # Ler e verificar usando csv.DictReader (lida com escaping automático)
    with open(output_file, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    # Verificar que dados foram preservados corretamente
    assert len(rows) == 1
    assert rows[0]["Razão Social"] == 'Empresa "Teste", LTDA'
    assert "Linha 1" in rows[0]["Observações"]


@patch("src.modules.clientes.export.is_xlsx_available")
@patch("tkinter.filedialog.asksaveasfilename")
def test_export_xlsx_option_only_if_available(mock_dialog, mock_is_xlsx, root, mock_vm, sample_clients):
    """FASE 3.5: Opção XLSX aparece apenas se openpyxl disponível."""
    frame = ClientesV2Frame(root)

    # Adicionar dados
    for client in sample_clients:
        iid = f"I{client.id:03d}"
        frame._row_data_map[iid] = client

    # Simular openpyxl NÃO disponível
    mock_is_xlsx.return_value = False
    mock_dialog.return_value = ""  # Cancelar

    # Chamar export
    frame._on_export()

    # Verificar que dialog foi chamado
    assert mock_dialog.called

    # Verificar filetypes passados ao dialog
    call_kwargs = mock_dialog.call_args[1]
    filetypes = call_kwargs["filetypes"]

    # Deve ter apenas CSV
    assert len(filetypes) == 1
    assert "CSV" in filetypes[0][0]

    # Agora simular openpyxl disponível
    mock_is_xlsx.return_value = True
    mock_dialog.reset_mock()

    frame._on_export()

    call_kwargs = mock_dialog.call_args[1]
    filetypes = call_kwargs["filetypes"]

    # Deve ter CSV e XLSX
    assert len(filetypes) == 2
    assert "CSV" in filetypes[0][0]
    assert "Excel" in filetypes[1][0] or "XLSX" in filetypes[1][0]
