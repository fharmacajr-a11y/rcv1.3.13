"""Testes unitários para exportação de clientes (CSV/XLSX).

Testa as funções headless de exportação sem abrir UI real.
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import TYPE_CHECKING

import pytest

if TYPE_CHECKING:
    pass


@pytest.fixture
def sample_rows():
    """Fixture com dados de clientes de exemplo para testes."""
    from src.modules.clientes.core.viewmodel import ClienteRow

    return [
        ClienteRow(
            id="1",
            razao_social="Empresa A Ltda",
            cnpj="12.345.678/0001-90",
            nome="Empresa A",
            whatsapp="11987654321",
            observacoes="Cliente ativo",
            status="Ativo",
            ultima_alteracao="2025-12-25 10:00:00",
        ),
        ClienteRow(
            id="2",
            razao_social="Empresa B S.A.",
            cnpj="98.765.432/0001-10",
            nome="Empresa B",
            whatsapp="11912345678",
            observacoes="[Pendente] Aguardando documentos",
            status="Pendente",
            ultima_alteracao="2025-12-24 15:30:00",
        ),
        ClienteRow(
            id="3",
            razao_social="Empresa C Comércio",
            cnpj="11.222.333/0001-44",
            nome="Loja C",
            whatsapp="",
            observacoes="",
            status="Inativo",
            ultima_alteracao="2025-12-20 08:15:00",
        ),
    ]


# ------------------------------------------------------------------ #
# Testes de exportação CSV
# ------------------------------------------------------------------ #


def test_export_csv_creates_file_with_headers(sample_rows, tmp_path):
    """Testa se exportação CSV cria arquivo com cabeçalhos corretos."""
    from src.modules.clientes.core.export import CSV_HEADERS, export_clients_to_csv

    output_file = tmp_path / "test_export.csv"

    # Executar exportação
    export_clients_to_csv(sample_rows, output_file)

    # Verificar que arquivo foi criado
    assert output_file.exists()

    # Ler arquivo e verificar cabeçalhos
    with open(output_file, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames

        # Verificar que todos os cabeçalhos esperados estão presentes
        assert headers is not None
        expected_headers = list(CSV_HEADERS.values())
        assert list(headers) == expected_headers


def test_export_csv_contains_correct_data(sample_rows, tmp_path):
    """Testa se dados exportados correspondem aos dados originais."""
    from src.modules.clientes.core.export import export_clients_to_csv

    output_file = tmp_path / "test_export.csv"

    # Executar exportação
    export_clients_to_csv(sample_rows, output_file)

    # Ler arquivo e verificar dados
    with open(output_file, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        rows = list(reader)

        # Verificar quantidade de linhas
        assert len(rows) == 3

        # Verificar primeira linha
        assert rows[0]["ID"] == "1"
        assert rows[0]["Razão Social"] == "Empresa A Ltda"
        assert rows[0]["CNPJ"] == "12.345.678/0001-90"
        assert rows[0]["Status"] == "Ativo"

        # Verificar segunda linha (com status em observações)
        assert rows[1]["ID"] == "2"
        assert rows[1]["Status"] == "Pendente"
        assert "[Pendente]" in rows[1]["Observações"]

        # Verificar terceira linha (campos vazios)
        assert rows[2]["ID"] == "3"
        assert rows[2]["WhatsApp"] == ""
        assert rows[2]["Observações"] == ""


def test_export_csv_handles_special_characters(tmp_path):
    """Testa se CSV lida corretamente com caracteres especiais e acentos."""
    from src.modules.clientes.core.export import export_clients_to_csv
    from src.modules.clientes.core.viewmodel import ClienteRow

    rows = [
        ClienteRow(
            id="1",
            razao_social="Açúcar & Café Ltda",
            cnpj="12.345.678/0001-90",
            nome="Café™",
            whatsapp="11987654321",
            observacoes="Observação com acentuação: é, ó, ã, ç",
            status="Ação",
            ultima_alteracao="2025-12-25 10:00:00",
        ),
    ]

    output_file = tmp_path / "test_special_chars.csv"

    # Executar exportação
    export_clients_to_csv(rows, output_file)

    # Ler arquivo e verificar caracteres especiais
    with open(output_file, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        data = list(reader)

        assert data[0]["Razão Social"] == "Açúcar & Café Ltda"
        assert data[0]["Nome Fantasia"] == "Café™"
        assert "acentuação" in data[0]["Observações"]


def test_export_csv_empty_list(tmp_path):
    """Testa exportação com lista vazia."""
    from src.modules.clientes.core.export import export_clients_to_csv

    output_file = tmp_path / "test_empty.csv"

    # Executar exportação com lista vazia
    export_clients_to_csv([], output_file)

    # Verificar que arquivo foi criado
    assert output_file.exists()

    # Verificar que contém apenas cabeçalho
    with open(output_file, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        assert len(rows) == 0


def test_export_csv_raises_on_invalid_path():
    """Testa se IOError é lançado para caminho inválido."""
    from src.modules.clientes.core.export import export_clients_to_csv

    invalid_path = Path("/caminho/inexistente/arquivo.csv")

    with pytest.raises(IOError):
        export_clients_to_csv([], invalid_path)


# ------------------------------------------------------------------ #
# Testes de exportação XLSX
# ------------------------------------------------------------------ #


def test_export_xlsx_requires_openpyxl(sample_rows, tmp_path):
    """Testa se exportação XLSX requer openpyxl instalado."""
    from src.modules.clientes.core.export import export_clients_to_xlsx, is_xlsx_available

    output_file = tmp_path / "test_export.xlsx"

    if not is_xlsx_available():
        # Se openpyxl não está instalado, deve lançar ImportError
        with pytest.raises(ImportError, match="openpyxl"):
            export_clients_to_xlsx(sample_rows, output_file)
    else:
        # Se openpyxl está instalado, deve funcionar
        export_clients_to_xlsx(sample_rows, output_file)
        assert output_file.exists()


def test_export_xlsx_creates_file_with_data(sample_rows, tmp_path):
    """Testa se exportação XLSX cria arquivo com dados corretos."""
    from src.modules.clientes.core.export import is_xlsx_available

    if not is_xlsx_available():
        pytest.skip("openpyxl não está instalado")

    from openpyxl import load_workbook

    from src.modules.clientes.core.export import export_clients_to_xlsx

    output_file = tmp_path / "test_export.xlsx"

    # Executar exportação
    export_clients_to_xlsx(sample_rows, output_file)

    # Verificar que arquivo foi criado
    assert output_file.exists()

    # Ler arquivo com openpyxl
    wb = load_workbook(output_file)
    ws = wb.active

    # Verificar cabeçalhos (primeira linha)
    assert ws is not None
    assert ws.cell(row=1, column=1).value == "ID"
    assert ws.cell(row=1, column=2).value == "Razão Social"
    assert ws.cell(row=1, column=3).value == "CNPJ"

    # Verificar dados (segunda linha)
    assert ws.cell(row=2, column=1).value == "1"
    assert ws.cell(row=2, column=2).value == "Empresa A Ltda"
    assert ws.cell(row=2, column=3).value == "12.345.678/0001-90"

    # Verificar quantidade de linhas (cabeçalho + 3 dados)
    assert ws.max_row == 4


def test_export_xlsx_handles_empty_list(tmp_path):
    """Testa exportação XLSX com lista vazia."""
    from src.modules.clientes.core.export import is_xlsx_available

    if not is_xlsx_available():
        pytest.skip("openpyxl não está instalado")

    from openpyxl import load_workbook

    from src.modules.clientes.core.export import export_clients_to_xlsx

    output_file = tmp_path / "test_empty.xlsx"

    # Executar exportação
    export_clients_to_xlsx([], output_file)

    # Verificar que arquivo foi criado
    assert output_file.exists()

    # Ler arquivo
    wb = load_workbook(output_file)
    ws = wb.active

    # Verificar que contém apenas cabeçalho
    assert ws is not None
    assert ws.max_row == 1  # Apenas linha de cabeçalho


# ------------------------------------------------------------------ #
# Testes de utilidades
# ------------------------------------------------------------------ #


def test_is_xlsx_available():
    """Testa detecção de disponibilidade de openpyxl."""
    from src.modules.clientes.core.export import is_xlsx_available

    # Apenas verifica que a função não falha
    result = is_xlsx_available()
    assert isinstance(result, bool)


# ------------------------------------------------------------------ #
# Testes de integração com viewmodel
# ------------------------------------------------------------------ #


def test_viewmodel_export_batch_validates_selection(monkeypatch):
    """Testa se export_batch valida seleção vazia."""
    from src.modules.clientes.core.viewmodel import ClientesViewModel

    # Mock messagebox
    warning_called = []

    def mock_showwarning(title, message):
        warning_called.append((title, message))

    monkeypatch.setattr("tkinter.messagebox.showwarning", mock_showwarning)

    # Mock cloud_guardrails (permitir operação)
    monkeypatch.setattr(
        "src.utils.helpers.cloud_guardrails.check_cloud_only_block",
        lambda operation_name: False,
    )

    # Criar viewmodel
    vm = ClientesViewModel()

    # Tentar exportar sem seleção
    vm.export_clientes_batch([])

    # Verificar que mostrou aviso
    assert len(warning_called) == 1
    assert "Nenhum cliente selecionado" in warning_called[0][1]


def test_viewmodel_export_batch_respects_cloud_only(monkeypatch):
    """Testa se export_batch respeita modo cloud-only."""
    from src.modules.clientes.core.viewmodel import ClientesViewModel

    # Mock cloud_guardrails (bloquear operação)
    cloud_block_called = []

    def mock_check_cloud_only_block(operation_name):
        cloud_block_called.append(operation_name)
        return True  # Bloquear

    monkeypatch.setattr(
        "src.utils.helpers.cloud_guardrails.check_cloud_only_block",
        mock_check_cloud_only_block,
    )

    # Criar viewmodel
    vm = ClientesViewModel()

    # Tentar exportar
    vm.export_clientes_batch(["1", "2"])

    # Verificar que cloud_only foi chamado e bloqueou
    assert len(cloud_block_called) == 1
    assert "Exportação de clientes" in cloud_block_called[0]
